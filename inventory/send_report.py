#!/usr/bin/env python3
"""Send pipeline report email with Excel attachment via AWS SES."""

import logging
import os
from datetime import datetime, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO

import boto3
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from inventory_diff import InventoryDiff

log = logging.getLogger(__name__)

SENDER = "admin@purpleblock.ai"
RECIPIENTS = [
    "shivam@purpleblock.ai",
    "ishanbhutoria@purpleblock.ai",
]
AWS_REGION = "us-east-1"

# Excel styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
GREEN_FILL = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
RED_FILL = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")
BLUE_FILL = PatternFill(start_color="BEE3F8", end_color="BEE3F8", fill_type="solid")


def _style_header_row(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _apply_border(ws) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def build_excel(diffs: list[InventoryDiff]) -> bytes:
    """Build Excel workbook with full inventory + changes per warehouse."""
    wb = Workbook()
    wb.remove(wb.active)

    for diff in diffs:
        # --- Full Inventory sheet (all products) ---
        sheet_name = f"{diff.warehouse}"[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["EAN", "SKU", "Name", "Price", "Stock", "Cost", "Cost + Tax"])
        _style_header_row(ws, 7)

        for item in diff.all_items:
            ws.append([
                item.get("ean", ""),
                item.get("sku", ""),
                item.get("name", ""),
                item.get("price", 0),
                item.get("stock", 0),
                item.get("cost", ""),
                item.get("costWithTax", ""),
            ])

        _apply_border(ws)
        _auto_width(ws)

        # --- Changes sheet (price + stock changes, new, removed) ---
        sheet_name = f"{diff.warehouse} - Changes"[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["EAN", "Name", "Type", "Old Value", "New Value", "Change"])
        _style_header_row(ws, 6)

        has_changes = diff.price_changes or diff.stock_changes or diff.new_items or diff.removed_items

        for item in diff.price_changes:
            change = item["new_price"] - item["old_price"]
            row_idx = ws.max_row + 1
            ws.append([item["ean"], item["name"], "Price", item["old_price"], item["new_price"], change])
            fill = RED_FILL if change > 0 else GREEN_FILL
            ws.cell(row=row_idx, column=6).fill = fill

        for item in diff.stock_changes:
            change = item["new_stock"] - item["old_stock"]
            row_idx = ws.max_row + 1
            ws.append([item["ean"], item["name"], "Stock", item["old_stock"], item["new_stock"], change])
            fill = GREEN_FILL if change > 0 else RED_FILL
            ws.cell(row=row_idx, column=6).fill = fill

        for item in diff.new_items:
            row_idx = ws.max_row + 1
            ws.append([item["ean"], item["name"], "New Item", "", item.get("price", 0), ""])
            ws.cell(row=row_idx, column=3).fill = GREEN_FILL

        for item in diff.removed_items:
            row_idx = ws.max_row + 1
            ws.append([item["ean"], item["name"], "Removed", item.get("price", 0), "", ""])
            ws.cell(row=row_idx, column=3).fill = RED_FILL

        if not has_changes:
            ws.append(["No changes detected", "", "", "", "", ""])

        _apply_border(ws)
        _auto_width(ws)

    # --- Summary sheet (first tab) ---
    ws = wb.create_sheet(title="Summary", index=0)
    ws.append(["Warehouse", "Total Items", "Price Changes", "Stock Changes", "New Items", "Removed Items"])
    _style_header_row(ws, 6)

    for diff in diffs:
        ws.append([
            diff.warehouse,
            diff.total_new,
            len(diff.price_changes),
            len(diff.stock_changes),
            len(diff.new_items),
            len(diff.removed_items),
        ])

    _apply_border(ws)
    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── HTML email (unchanged) ──────────────────────────────────────────

def _fmt_number(n: int) -> str:
    return f"{n:,}"


def _price_arrow(old: int, new: int) -> str:
    if new > old:
        return f"${_fmt_number(old)} &rarr; <span style='color:#e53e3e'>${_fmt_number(new)} (+{_fmt_number(new - old)})</span>"
    return f"${_fmt_number(old)} &rarr; <span style='color:#38a169'>${_fmt_number(new)} ({_fmt_number(new - old)})</span>"


def _stock_arrow(old: int, new: int) -> str:
    if new > old:
        return f"{_fmt_number(old)} &rarr; <span style='color:#38a169'>{_fmt_number(new)} (+{_fmt_number(new - old)})</span>"
    return f"{_fmt_number(old)} &rarr; <span style='color:#e53e3e'>{_fmt_number(new)} ({_fmt_number(new - old)})</span>"


def _build_warehouse_section(diff: InventoryDiff) -> str:
    has_changes = (
        diff.new_items or diff.removed_items
        or diff.price_changes or diff.stock_changes
    )

    html = f"""
    <div style="margin-bottom:24px;border:1px solid #e2e8f0;border-radius:8px;overflow:hidden">
      <div style="background:#2d3748;color:#fff;padding:12px 16px;font-size:16px;font-weight:600">
        {diff.warehouse}
        <span style="font-weight:400;font-size:13px;opacity:0.8">
          &mdash; {_fmt_number(diff.total_new)} items
        </span>
      </div>
      <div style="padding:16px">
    """

    if not has_changes:
        html += '<p style="color:#718096;margin:0">No changes detected.</p>'
        html += "</div></div>"
        return html

    badges = []
    if diff.new_items:
        badges.append(f'<span style="background:#c6f6d5;color:#276749;padding:2px 8px;border-radius:4px;font-size:13px">+{len(diff.new_items)} new</span>')
    if diff.removed_items:
        badges.append(f'<span style="background:#fed7d7;color:#9b2c2c;padding:2px 8px;border-radius:4px;font-size:13px">-{len(diff.removed_items)} removed</span>')
    if diff.price_changes:
        badges.append(f'<span style="background:#fefcbf;color:#975a16;padding:2px 8px;border-radius:4px;font-size:13px">{len(diff.price_changes)} price changes</span>')
    if diff.stock_changes:
        badges.append(f'<span style="background:#bee3f8;color:#2a4365;padding:2px 8px;border-radius:4px;font-size:13px">{len(diff.stock_changes)} stock changes</span>')

    html += f'<div style="margin-bottom:12px">{" ".join(badges)}</div>'

    table_style = 'style="width:100%;border-collapse:collapse;font-size:13px;margin-bottom:16px"'
    th_style = 'style="text-align:left;padding:6px 8px;border-bottom:2px solid #e2e8f0;color:#4a5568;font-size:12px;text-transform:uppercase"'
    td_style = 'style="padding:6px 8px;border-bottom:1px solid #edf2f7"'

    if diff.price_changes:
        html += f'<h4 style="margin:0 0 6px;font-size:14px;color:#2d3748">Price Changes</h4>'
        html += f"<table {table_style}><thead><tr>"
        html += f"<th {th_style}>EAN</th><th {th_style}>Name</th><th {th_style}>Price Change</th>"
        html += "</tr></thead><tbody>"
        for item in diff.price_changes[:50]:
            html += f"<tr><td {td_style}>{item['ean']}</td><td {td_style}>{item['name']}</td>"
            html += f"<td {td_style}>{_price_arrow(item['old_price'], item['new_price'])}</td></tr>"
        if len(diff.price_changes) > 50:
            html += f'<tr><td {td_style} colspan="3" style="color:#718096;font-style:italic">...and {len(diff.price_changes) - 50} more (see attached Excel)</td></tr>'
        html += "</tbody></table>"

    if diff.stock_changes:
        html += f'<h4 style="margin:0 0 6px;font-size:14px;color:#2d3748">Stock Changes</h4>'
        html += f"<table {table_style}><thead><tr>"
        html += f"<th {th_style}>EAN</th><th {th_style}>Name</th><th {th_style}>Stock Change</th>"
        html += "</tr></thead><tbody>"
        for item in diff.stock_changes[:50]:
            html += f"<tr><td {td_style}>{item['ean']}</td><td {td_style}>{item['name']}</td>"
            html += f"<td {td_style}>{_stock_arrow(item['old_stock'], item['new_stock'])}</td></tr>"
        if len(diff.stock_changes) > 50:
            html += f'<tr><td {td_style} colspan="3" style="color:#718096;font-style:italic">...and {len(diff.stock_changes) - 50} more (see attached Excel)</td></tr>'
        html += "</tbody></table>"

    if diff.new_items:
        html += f'<h4 style="margin:0 0 6px;font-size:14px;color:#2d3748">New Items</h4>'
        html += f"<table {table_style}><thead><tr>"
        html += f"<th {th_style}>EAN</th><th {th_style}>Name</th><th {th_style}>Price</th><th {th_style}>Stock</th>"
        html += "</tr></thead><tbody>"
        for item in diff.new_items[:30]:
            html += f"<tr><td {td_style}>{item['ean']}</td><td {td_style}>{item['name']}</td>"
            html += f"<td {td_style}>${_fmt_number(item['price'])}</td><td {td_style}>{_fmt_number(item['stock'])}</td></tr>"
        if len(diff.new_items) > 30:
            html += f'<tr><td {td_style} colspan="4" style="color:#718096;font-style:italic">...and {len(diff.new_items) - 30} more (see attached Excel)</td></tr>'
        html += "</tbody></table>"

    if diff.removed_items:
        html += f'<h4 style="margin:0 0 6px;font-size:14px;color:#2d3748">Removed Items</h4>'
        html += f"<table {table_style}><thead><tr>"
        html += f"<th {th_style}>EAN</th><th {th_style}>Name</th><th {th_style}>Last Price</th><th {th_style}>Last Stock</th>"
        html += "</tr></thead><tbody>"
        for item in diff.removed_items[:30]:
            html += f"<tr><td {td_style}>{item['ean']}</td><td {td_style}>{item['name']}</td>"
            html += f"<td {td_style}>${_fmt_number(item['price'])}</td><td {td_style}>{_fmt_number(item['stock'])}</td></tr>"
        if len(diff.removed_items) > 30:
            html += f'<tr><td {td_style} colspan="4" style="color:#718096;font-style:italic">...and {len(diff.removed_items) - 30} more (see attached Excel)</td></tr>'
        html += "</tbody></table>"

    html += "</div></div>"
    return html


def build_report_html(
    diffs: list[InventoryDiff],
    failures: list[tuple[str, str]],
) -> str:
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    total_price_changes = sum(len(d.price_changes) for d in diffs)
    total_stock_changes = sum(len(d.stock_changes) for d in diffs)
    total_new = sum(len(d.new_items) for d in diffs)
    total_removed = sum(len(d.removed_items) for d in diffs)

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f7fafc;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif">
<div style="max-width:700px;margin:0 auto;padding:24px">
  <div style="background:#1a202c;color:#fff;padding:20px 24px;border-radius:8px 8px 0 0">
    <h1 style="margin:0;font-size:20px;font-weight:600">Kamal Inventory Pipeline Report</h1>
    <p style="margin:4px 0 0;font-size:13px;opacity:0.7">{now}</p>
  </div>
  <div style="background:#fff;padding:24px;border-radius:0 0 8px 8px;border:1px solid #e2e8f0;border-top:0">
"""

    if failures:
        html += '<div style="background:#fed7d7;border:1px solid #fc8181;border-radius:6px;padding:12px 16px;margin-bottom:16px">'
        html += '<strong style="color:#9b2c2c">Pipeline Failures</strong><ul style="margin:8px 0 0;padding-left:20px">'
        for step_name, error in failures:
            html += f'<li style="color:#9b2c2c"><strong>{step_name}:</strong> {error}</li>'
        html += "</ul></div>"

    html += f"""
    <div style="display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap">
      <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:6px;padding:10px 16px;flex:1;min-width:120px;text-align:center">
        <div style="font-size:24px;font-weight:700;color:#2d3748">{len(diffs)}</div>
        <div style="font-size:12px;color:#718096">Warehouses</div>
      </div>
      <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:6px;padding:10px 16px;flex:1;min-width:120px;text-align:center">
        <div style="font-size:24px;font-weight:700;color:#d69e2e">{total_price_changes}</div>
        <div style="font-size:12px;color:#718096">Price Changes</div>
      </div>
      <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:6px;padding:10px 16px;flex:1;min-width:120px;text-align:center">
        <div style="font-size:24px;font-weight:700;color:#3182ce">{total_stock_changes}</div>
        <div style="font-size:12px;color:#718096">Stock Changes</div>
      </div>
      <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:6px;padding:10px 16px;flex:1;min-width:120px;text-align:center">
        <div style="font-size:24px;font-weight:700;color:#38a169">{total_new}</div>
        <div style="font-size:12px;color:#718096">New Items</div>
      </div>
      <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:6px;padding:10px 16px;flex:1;min-width:120px;text-align:center">
        <div style="font-size:24px;font-weight:700;color:#e53e3e">{total_removed}</div>
        <div style="font-size:12px;color:#718096">Removed</div>
      </div>
    </div>
"""

    for diff in diffs:
        html += _build_warehouse_section(diff)

    html += """
    <p style="margin:16px 0 0;font-size:12px;color:#a0aec0;text-align:center">
      Full data attached as Excel &bull; Automated report from Kamal Inventory Pipeline
    </p>
  </div>
</div>
</body></html>"""
    return html


# ── Send ────────────────────────────────────────────────────────────

def send_report(
    diffs: list[InventoryDiff],
    failures: list[tuple[str, str]],
) -> None:
    """Build HTML + Excel and send via SES raw email."""
    html = build_report_html(diffs, failures)
    excel_bytes = build_excel(diffs)

    total_changes = sum(
        len(d.price_changes) + len(d.stock_changes) + len(d.new_items) + len(d.removed_items)
        for d in diffs
    )

    subject = f"Kamal Pipeline Report \u2014 {total_changes} changes"
    if failures:
        subject = f"[FAILURES] {subject}"

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"inventory-diff-{today}.xlsx"

    ses = boto3.client("ses", region_name=AWS_REGION)

    for recipient in RECIPIENTS:
        try:
            msg = MIMEMultipart("mixed")
            msg["Subject"] = subject
            msg["From"] = SENDER
            msg["To"] = recipient

            # HTML body
            body_part = MIMEText(html, "html", "utf-8")
            msg.attach(body_part)

            # Excel attachment
            attachment = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            attachment.set_payload(excel_bytes)
            encoders.encode_base64(attachment)
            attachment.add_header("Content-Disposition", "attachment", filename=filename)
            msg.attach(attachment)

            resp = ses.send_raw_email(
                Source=SENDER,
                Destinations=[recipient],
                RawMessage={"Data": msg.as_string()},
            )
            log.info("Report email sent to %s (MessageId: %s)", recipient, resp["MessageId"])
        except Exception as e:
            log.warning("Failed to send report email to %s: %s", recipient, e)
