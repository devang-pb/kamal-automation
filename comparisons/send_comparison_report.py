#!/usr/bin/env python3
"""Send comparison pipeline report email with full Excel via AWS SES."""

import csv
import logging
import os
from datetime import datetime, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO

import boto3
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

SENDER = "admin@purpleblock.ai"
RECIPIENTS = [
    "shivam@purpleblock.ai",
    "ishanbhutoria@purpleblock.ai",
]
AWS_REGION = "us-east-1"

MERGED_CSV = os.path.join(os.environ.get("OUTPUT_DIR", "output"), "merged_comparisons.csv")

# Excel styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
GREEN_FILL = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
RED_FILL = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")


def _parse_price(val):
    if not val or not str(val).strip():
        return None
    try:
        return float(str(val).strip().replace(",", ""))
    except ValueError:
        return None


def _parse_gap(val):
    if not val or not str(val).strip():
        return None
    try:
        return float(str(val).strip().rstrip("%"))
    except ValueError:
        return None


def load_merged_csv() -> list[dict]:
    """Load merged_comparisons.csv and return rows as dicts."""
    if not os.path.exists(MERGED_CSV):
        log.error("merged_comparisons.csv not found at %s", MERGED_CSV)
        return []
    with open(MERGED_CSV, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def build_excel(rows: list[dict]) -> bytes:
    """Build Excel from merged comparison data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Data"

    if not rows:
        ws.append(["No comparison data available"])
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Write headers
    headers = list(rows[0].keys())
    ws.append(headers)
    num_cols = len(headers)

    # Style header row
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Find column indices
    gap_col = None
    my_price_col = None
    cheapest_col = None
    for i, h in enumerate(headers):
        if h == "Price Gap %":
            gap_col = i + 1
        elif h == "My Price":
            my_price_col = i + 1
        elif h == "Cheapest Price":
            cheapest_col = i + 1

    # Write data rows
    for row_data in rows:
        values = list(row_data.values())
        row_idx = ws.max_row + 1
        ws.append(values)

        # Color the price gap column
        if gap_col:
            gap_val = _parse_gap(row_data.get("Price Gap %"))
            if gap_val is not None:
                cell = ws.cell(row=row_idx, column=gap_col)
                if gap_val < 0:
                    cell.fill = GREEN_FILL  # We're cheaper
                elif gap_val > 0:
                    cell.fill = RED_FILL  # We're more expensive

    # Apply borders to data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=num_cols):
        for cell in row:
            cell.border = THIN_BORDER

    # Auto-width
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # Freeze header row + first 2 columns (barcode, name)
    ws.freeze_panes = "C2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_html(rows: list[dict], failures: list[tuple[str, str]]) -> str:
    """Build summary HTML email for comparison pipeline."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    total_products = len(rows)

    # Compute stats
    cheaper_count = 0
    expensive_count = 0
    no_data_count = 0
    for row in rows:
        gap = _parse_gap(row.get("Price Gap %"))
        if gap is None:
            no_data_count += 1
        elif gap < 0:
            cheaper_count += 1
        elif gap > 0:
            expensive_count += 1

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f7fafc;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif">
<div style="max-width:700px;margin:0 auto;padding:24px">
  <div style="background:#1a202c;color:#fff;padding:20px 24px;border-radius:8px 8px 0 0">
    <h1 style="margin:0;font-size:20px;font-weight:600">Kamal Comparison Pipeline Report</h1>
    <p style="margin:4px 0 0;font-size:13px;opacity:0.7">{now}</p>
  </div>
  <div style="background:#fff;padding:24px;border-radius:0 0 8px 8px;border:1px solid #e2e8f0;border-top:0">
"""

    if failures:
        html += '<div style="background:#fed7d7;border:1px solid #fc8181;border-radius:6px;padding:12px 16px;margin-bottom:16px">'
        html += '<strong style="color:#9b2c2c">Pipeline Failures</strong><ul style="margin:8px 0 0;padding-left:20px">'
        for step_name, error in failures:
            html += f'<li style="color:#9b2c2c"><strong>{step_name}:</strong> {error}</li>'
        html += "</ul></div>"

    html += f"""
    <div style="display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap">
      <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:6px;padding:10px 16px;flex:1;min-width:120px;text-align:center">
        <div style="font-size:24px;font-weight:700;color:#2d3748">{total_products:,}</div>
        <div style="font-size:12px;color:#718096">Products Compared</div>
      </div>
      <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:6px;padding:10px 16px;flex:1;min-width:120px;text-align:center">
        <div style="font-size:24px;font-weight:700;color:#38a169">{cheaper_count:,}</div>
        <div style="font-size:12px;color:#718096">We're Cheaper</div>
      </div>
      <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:6px;padding:10px 16px;flex:1;min-width:120px;text-align:center">
        <div style="font-size:24px;font-weight:700;color:#e53e3e">{expensive_count:,}</div>
        <div style="font-size:12px;color:#718096">More Expensive</div>
      </div>
      <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:6px;padding:10px 16px;flex:1;min-width:120px;text-align:center">
        <div style="font-size:24px;font-weight:700;color:#718096">{no_data_count:,}</div>
        <div style="font-size:12px;color:#718096">No Competitor Data</div>
      </div>
    </div>
"""

    # Top 15 where we're most expensive
    expensive = []
    for row in rows:
        gap = _parse_gap(row.get("Price Gap %"))
        if gap is not None and gap > 0:
            expensive.append((gap, row))
    expensive.sort(key=lambda x: x[0], reverse=True)

    if expensive:
        table_style = 'style="width:100%;border-collapse:collapse;font-size:13px;margin-bottom:16px"'
        th_style = 'style="text-align:left;padding:6px 8px;border-bottom:2px solid #e2e8f0;color:#4a5568;font-size:12px;text-transform:uppercase"'
        td_style = 'style="padding:6px 8px;border-bottom:1px solid #edf2f7"'

        html += '<h4 style="margin:0 0 6px;font-size:14px;color:#2d3748">Top Products Where We\'re Most Expensive</h4>'
        html += f"<table {table_style}><thead><tr>"
        html += f"<th {th_style}>EAN</th><th {th_style}>Name</th><th {th_style}>Our Price</th><th {th_style}>Cheapest</th><th {th_style}>Gap</th><th {th_style}>Cheapest Site</th>"
        html += "</tr></thead><tbody>"
        for gap, row in expensive[:15]:
            html += f"<tr><td {td_style}>{row.get('Bar Code','')}</td>"
            html += f"<td {td_style}>{row.get('Name','')}</td>"
            html += f"<td {td_style}>${row.get('My Price','')}</td>"
            html += f"<td {td_style}>${row.get('Cheapest Price','')}</td>"
            html += f"<td {td_style}><span style='color:#e53e3e;font-weight:600'>+{gap:.1f}%</span></td>"
            html += f"<td {td_style}>{row.get('Cheapest Site','')}</td></tr>"
        html += "</tbody></table>"

    html += """
    <p style="margin:16px 0 0;font-size:12px;color:#a0aec0;text-align:center">
      Full data attached as Excel &bull; Automated report from Kamal Comparison Pipeline
    </p>
  </div>
</div>
</body></html>"""
    return html


def send_comparison_report(failures: list[tuple[str, str]] | None = None) -> None:
    """Load merged CSV, build Excel + HTML, send email via SES."""
    failures = failures or []
    rows = load_merged_csv()

    html = _build_html(rows, failures)
    excel_bytes = build_excel(rows)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    subject = f"Kamal Comparison Report \u2014 {len(rows):,} products"
    if failures:
        subject = f"[FAILURES] {subject}"
    filename = f"comparison-data-{today}.xlsx"

    ses = boto3.client("ses", region_name=AWS_REGION)

    for recipient in RECIPIENTS:
        try:
            msg = MIMEMultipart("mixed")
            msg["Subject"] = subject
            msg["From"] = SENDER
            msg["To"] = recipient

            body_part = MIMEText(html, "html", "utf-8")
            msg.attach(body_part)

            attachment = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            attachment.set_payload(excel_bytes)
            encoders.encode_base64(attachment)
            attachment.add_header("Content-Disposition", "attachment", filename=filename)
            msg.attach(attachment)

            resp = ses.send_raw_email(
                Source=SENDER,
                Destinations=[recipient],
                RawMessage={"Data": msg.as_string()},
            )
            log.info("Comparison report sent to %s (MessageId: %s)", recipient, resp["MessageId"])
        except Exception as e:
            log.warning("Failed to send comparison report to %s: %s", recipient, e)
